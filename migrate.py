import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

XLSX_PATH = "/Users/jacksonsharkey/Desktop/CommuneUSA.xlsx"
OUTPUT_DIR = Path(__file__).parent / "output"

SHEET_MAP = {
    "County Officials": "county_officials.json",
    "WA Officials": "wa_officials.json",
    "State Legislature": "state_legislature.json",
    "Federal Legislators": "federal_legislators.json",
}

VOTING_SHEETS = ["Voting Record", "Voting Record1"]

# Maps cleaned Excel column names → voting_records schema fields
VOTING_COL_MAP = {
    "official_name":      "official_name",
    "bill_motion":        "bill_name",
    "topic_category":     "topic_category",
    "date":               "vote_date",
    "their_vote":         "vote_cast",
    "result":             "result",
    "constituent_impact": "constituent_impact",
    "source_url":         "source_url",
}

CAMPAIGN_FINANCE_SHEETS = ["Campaign Funding", "Campaign Funding1"]

ELECTION_SHEETS = ["Upcoming Elections", "Upcoming Elections1"]

# Placeholder patterns that are not real candidate names
_PLACEHOLDER_RE = re.compile(
    r"^(tbd|unknown|various|multiple(\s+tbd|\s+candidates?)?|open\s+field.*|"
    r"challengers?.*|others?.*|full\s+candidate.*|full\s+list.*|n/a|none|—|-|\+|"
    r"and|\d+\s+seats?\s+(open|up).*|all\s+seats?.*|several\s+open.*|"
    r"incumbents?\s+likely.*|top\s+2.*)$",
    re.IGNORECASE
)

# Per-sheet column maps: cleaned Excel header → campaign_finance schema field.
# Sheet 1 uses older column names; sheet 2 uses the canonical schema names.
CAMPAIGN_COL_MAPS = {
    "Campaign Funding": {
        "official_name":      "official_name",
        "top_donor_category": "donor_name",
        "donor_type":         "donor_type",
        "amount":             "amount",
        "election_year":      "election_cycle",
        "industry_sector":    "industry_sector",
        "pdc_fec_filing":     "source_url",
    },
    "Campaign Funding1": {
        "official_name":  "official_name",
        "donor_name":     "donor_name",
        "donor_type":     "donor_type",
        "amount":         "amount",
        "election_cycle": "election_cycle",
        "date":           "donation_date",
        "industry_sector": "industry_sector",
        "source_url":     "source_url",
    },
}


def clean_header(raw):
    if raw is None:
        return None
    text = str(raw)
    text = "".join(c for c in text if unicodedata.category(c) != "So" and ord(c) < 128)
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    return text or None


def serialize_value(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, float):
        return int(val) if val == int(val) else val
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def is_section_row(row_values, headers):
    first = row_values[0]
    if first is not None and "▶" in str(first):
        return True
    populated = sum(1 for v in row_values if v is not None and str(v).strip())
    threshold = max(1, len(headers) * 0.30)
    return populated < threshold


def find_sheet(wb, name):
    target = name.lower()
    for sheet_name in wb.sheetnames:
        clean = re.sub(r"[^\x00-\x7F]", "", sheet_name).strip().lower()
        if target in clean:
            return wb[sheet_name]
    return None


def extract_records(ws):
    rows = list(ws.iter_rows(values_only=True))

    # Find header row: first row with more than one populated cell
    header_idx = None
    for i, row in enumerate(rows):
        populated = sum(1 for v in row if v is not None and str(v).strip())
        if populated > 1:
            header_idx = i
            break

    if header_idx is None:
        return []

    raw_headers = rows[header_idx]
    headers = [clean_header(h) for h in raw_headers]

    records = []
    for row in rows[header_idx + 1 :]:
        row_values = list(row)

        # Pad or trim to match header length
        while len(row_values) < len(headers):
            row_values.append(None)
        row_values = row_values[: len(headers)]

        if is_section_row(row_values, headers):
            continue

        # Skip entirely empty rows
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        record = {}
        for key, val in zip(headers, row_values):
            if key is None:
                continue
            record[key] = serialize_value(val)
        records.append(record)

    return records


def extract_voting_records(wb):
    all_records = []
    for sheet_key in VOTING_SHEETS:
        ws = find_sheet(wb, sheet_key)
        if ws is None:
            print(f"  WARN  sheet not found: {sheet_key}")
            continue
        raw = extract_records(ws)
        count = 0
        for rec in raw:
            mapped = {
                "official_name":    rec.get("official_name"),
                "official_id":      None,
                "bill_name":        rec.get("bill_motion"),
                "bill_description": None,
                "topic_category":   rec.get("topic_category"),
                "vote_date":        rec.get("date"),
                "vote_cast":        rec.get("their_vote"),
                "result":           rec.get("result"),
                "constituent_impact": rec.get("constituent_impact"),
                "source_url":       rec.get("source_url"),
            }
            if not mapped["official_name"] and not mapped["bill_name"]:
                continue
            all_records.append(mapped)
            count += 1
        print(f"  OK    {sheet_key} → {count} records extracted")
    return all_records


def parse_amount(val):
    """Return a numeric value from strings like '~$1,200,000' or plain numbers."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val) if isinstance(val, float) and val == int(val) else val
    s = str(val).replace("~", "").replace("$", "").replace(",", "").strip()
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, OverflowError):
        return None


def _is_placeholder_official(name) -> bool:
    """Return True for rows that are instructions or catch-all entries, not real officials."""
    if not name:
        return True
    n = str(name).strip().lower()
    return (
        n.startswith("n/a")
        or n.startswith("search")
        or "pdc.wa.gov" in n
        or n == ""
    )


def extract_campaign_finance(wb):
    all_records = []
    for sheet_key in CAMPAIGN_FINANCE_SHEETS:
        ws = find_sheet(wb, sheet_key)
        if ws is None:
            print(f"  WARN  sheet not found: {sheet_key}")
            continue
        col_map = CAMPAIGN_COL_MAPS[sheet_key]
        raw = extract_records(ws)
        count = 0
        for rec in raw:
            official_name = rec.get("official_name")
            if _is_placeholder_official(official_name):
                continue

            mapped = {
                "official_name":   official_name,
                "official_id":     None,
                "donor_name":      None,
                "donor_type":      None,
                "amount":          None,
                "election_cycle":  None,
                "donation_date":   None,
                "industry_sector": None,
                "source_url":      None,
                "filing_source":   None,
            }
            for excel_col, target_field in col_map.items():
                val = rec.get(excel_col)
                if target_field == "amount":
                    mapped["amount"] = parse_amount(val)
                else:
                    mapped[target_field] = val

            all_records.append(mapped)
            count += 1
        print(f"  OK    {sheet_key} → {count} records extracted")
    return all_records


def _infer_level_from_muni(raw: str) -> str:
    """Infer election level from Sheet 1's 'Municipality / Level' column."""
    s = (raw or "").strip().lower()
    if "county" in s:
        return "county"
    if s in ("state of washington", "washington state", "wa"):
        return "state"
    if "federal" in s:
        return "federal"
    return "city"


def _parse_level_explicit(raw: str) -> str:
    """Parse explicit level from Sheet 2's 'Gov. Level' column."""
    s = (raw or "").strip().lower()
    if s == "federal":
        return "federal"
    if s == "state":
        return "state"
    if s == "county":
        return "county"
    # "city/state" → treat as state since it covers multiple jurisdictions
    if "state" in s and "city" in s:
        return "state"
    return "city"


def _normalize_election_date(raw) -> Optional[str]:
    """Normalize election date to YYYY-MM-DD or YYYY-MM where possible."""
    if not raw:
        return None
    s = str(raw).strip()

    # "Aug 4, 2026 (Primary) / Nov 3 (General)" → extract General date
    if "/" in s and re.search(r"primary|general", s, re.IGNORECASE):
        after = s.split("/", 1)[1].strip()
        after = re.sub(r"\s*\([^)]*\)", "", after).strip()
        # Append year from the original string if the fragment has no year
        if not re.search(r"\b20\d{2}\b", after):
            year_m = re.search(r"\b(20\d{2})\b", s)
            if year_m:
                after = f"{after}, {year_m.group(1)}"
        s = after

    # Strip remaining parentheticals
    s = re.sub(r"\s*\([^)]*\)", "", s).strip()

    for fmt in ("%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    for fmt in ("%b %Y", "%B %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m")
        except ValueError:
            pass
    return s  # return as-is if unparseable


def _clean_name(raw) -> Optional[str]:
    """Strip parenthetical qualifiers from a person's name."""
    if not raw:
        return None
    clean = re.sub(r"\s*\([^)]*\)", "", str(raw)).strip()
    return clean or None


def _parse_candidate_names(raw) -> list[str]:
    """
    Extract clean candidate names from a messy 'Candidate(s) Known' or incumbent
    field. Handles comma-separated lists, 'vs.', and '+' delimiters. Returns []
    for pure placeholder values.
    """
    if not raw:
        return []
    s = str(raw).strip()
    if _PLACEHOLDER_RE.match(s):
        return []

    parts = re.split(r",\s*|\s+\+\s+|\s+vs\.?\s+", s)
    names = []
    for part in parts:
        clean = re.sub(r"\s*\([^)]*\)", "", part).strip()
        if not clean or _PLACEHOLDER_RE.match(clean):
            continue
        if re.search(
            r"\b(tbd|challengers?|others?|multiple|full\s+candidate|"
            r"sos\.wa\.gov|leg\.wa\.gov|ballotpedia|seattle\.gov)\b",
            clean, re.IGNORECASE,
        ):
            continue
        if len(clean.split()) > 5:  # too long to be a person's name
            continue
        names.append(clean)
    return names


def _election_key(office_name: str, election_date: Optional[str], location: str = "") -> str:
    return (
        f"{(location or '').strip().lower()}||"
        f"{(office_name or '').strip().lower()}||"
        f"{(election_date or '').strip()}"
    )


def extract_elections(wb):
    """
    Extract elections and candidates from both election sheets.

    Returns (elections, candidates) where:
      - Each election has a '_key' field used for candidate linking (not stored).
      - Each candidate has an 'election_key' referencing its election's '_key'.
    """
    elections: list[dict] = []
    candidates: list[dict] = []
    seen_keys: set[str] = set()

    # ── Sheet 1: "Municipality / Level" layout ────────────────────────────────
    ws1 = find_sheet(wb, "Upcoming Elections")
    if ws1 is None:
        print("  WARN  sheet not found: Upcoming Elections")
    else:
        raw1 = extract_records(ws1)
        count_e = count_c = 0

        for rec in raw1:
            office_name = rec.get("office_race")
            if not office_name:
                continue

            # Skip meta-rows (e.g. "Primary Election" with status "Primary")
            status = (rec.get("incumbent_status") or "").strip()
            if str(office_name).lower().startswith("primary") and "primary" in status.lower():
                continue

            muni_raw = str(rec.get("municipality_level") or "").strip()
            level = _infer_level_from_muni(muni_raw)
            election_date = _normalize_election_date(rec.get("election_date"))
            key = _election_key(office_name, election_date, muni_raw)

            if key in seen_keys:
                continue
            seen_keys.add(key)

            elections.append({
                "_key":                   key,
                "office_name":            str(office_name).strip(),
                "level":                  level,
                "municipality_level_raw": muni_raw,
                "election_date":          election_date,
                "primary_date":           _normalize_election_date(rec.get("primary_date")),
                "filing_deadline":        rec.get("filing_deadline"),
                "description":            rec.get("key_issues"),
                "source_url":             rec.get("source_url"),
            })
            count_e += 1

            # Incumbent candidate
            inc_raw = rec.get("incumbent") or ""
            inc_name = _clean_name(inc_raw)
            if (
                inc_name
                and inc_name.lower() not in ("various", "tbd", "none")
                and "open race" not in status.lower()
            ):
                candidates.append({
                    "election_key": key,
                    "name":         inc_name,
                    "is_incumbent": True,
                    "party":        None,
                    "website":      None,
                    "ballotpedia_url": None,
                })
                count_c += 1

            # Additional candidates from "Candidate(s) Known"
            for cand_name in _parse_candidate_names(rec.get("candidates_known")):
                # Skip if it matches the incumbent (already added)
                if inc_name and cand_name.lower() in inc_name.lower():
                    continue
                candidates.append({
                    "election_key": key,
                    "name":         cand_name,
                    "is_incumbent": False,
                    "party":        None,
                    "website":      None,
                    "ballotpedia_url": None,
                })
                count_c += 1

        print(f"  OK    Upcoming Elections → {count_e} elections, {count_c} candidates extracted")

    # ── Sheet 2: "Gov. Level" layout ──────────────────────────────────────────
    ws2 = find_sheet(wb, "Upcoming Elections1")
    if ws2 is None:
        print("  WARN  sheet not found: Upcoming Elections1")
    else:
        raw2 = extract_records(ws2)
        count_e = count_c = 0

        for rec in raw2:
            office_name = rec.get("office_race")
            if not office_name:
                continue

            status = (rec.get("status") or "").strip()
            if str(office_name).lower().startswith("primary") and "primary" in status.lower():
                continue

            gov_level_raw = str(rec.get("gov_level") or "").strip()
            level = _parse_level_explicit(gov_level_raw)
            election_date = _normalize_election_date(rec.get("election_date"))
            muni_context = "Seattle" if level == "city" else gov_level_raw
            key = _election_key(office_name, election_date, muni_context)

            if key in seen_keys:
                continue
            seen_keys.add(key)

            # Prefer source_url; fall back to campaign_website
            source_url = rec.get("source_url") or rec.get("campaign_website")

            elections.append({
                "_key":                   key,
                "office_name":            str(office_name).strip(),
                "level":                  level,
                "municipality_level_raw": muni_context,
                "election_date":          election_date,
                "primary_date":           None,
                "filing_deadline":        rec.get("filing_deadline"),
                "description":            rec.get("key_issues"),
                "source_url":             source_url,
            })
            count_e += 1

            # Incumbent(s) — Sheet 2 may have comma-separated names
            inc_raw = rec.get("incumbent") or ""
            inc_names = (
                _parse_candidate_names(inc_raw)
                if "," in inc_raw
                else ([_clean_name(inc_raw)] if _clean_name(inc_raw) else [])
            )
            for inc_name in inc_names:
                if not inc_name or inc_name.lower() in ("various", "tbd", "none"):
                    continue
                if "open race" in status.lower():
                    continue
                candidates.append({
                    "election_key": key,
                    "name":         inc_name,
                    "is_incumbent": True,
                    "party":        None,
                    "website":      None,
                    "ballotpedia_url": None,
                })
                count_c += 1

        print(f"  OK    Upcoming Elections1 → {count_e} elections, {count_c} candidates extracted")

    return elections, candidates


def _map_records(raw: list[dict], col_map: dict[str, str], required_key: str) -> list[dict]:
    """Map raw extracted records through col_map; skip rows missing required_key."""
    out = []
    for rec in raw:
        val = rec.get(required_key)
        if not val or not str(val).strip():
            continue
        out.append({target: rec.get(src) for src, target in col_map.items()})
    return out


def extract_school_boards(wb) -> list[dict]:
    ws = find_sheet(wb, "School Boards")
    if ws is None:
        print("  WARN  sheet not found: School Boards")
        return []
    raw = extract_records(ws)
    col_map = {
        "district_name":     "district_name",
        "county":            "county",
        "board_director_name": "director_name",
        "position_director": "position",
        "party_affiliation": "party_affiliation",
        "term_start":        "term_start",
        "term_end":          "term_end",
        "phone_email":       "phone",
        "official_website":  "website",
    }
    records = _map_records(raw, col_map, "board_director_name")
    print(f"  OK    School Boards → school_boards.json ({len(records)} records)")
    return records


def extract_state_agencies(wb) -> list[dict]:
    ws = find_sheet(wb, "State Agencies")
    if ws is None:
        print("  WARN  sheet not found: State Agencies")
        return []
    raw = extract_records(ws)
    col_map = {
        "category":          "category",
        "agency_department": "name",
        "abbreviation":      "abbreviation",
        "director_secretary": "director_name",
        "selection":         "selection_method",
        "budget_approx":     "budget",
        "employees":         "employees",
        "headquarters":      "headquarters",
        "phone":             "phone",
        "website":           "website",
        "mission_summary":   "mission_summary",
    }
    records = _map_records(raw, col_map, "agency_department")
    print(f"  OK    State Agencies → state_agencies.json ({len(records)} records)")
    return records


def extract_law_enforcement(wb) -> list[dict]:
    ws = find_sheet(wb, "Law Enforcement")
    if ws is None:
        print("  WARN  sheet not found: Law Enforcement")
        return []
    raw = extract_records(ws)
    col_map = {
        "agency_type":          "agency_type",
        "agency_name":          "name",
        "jurisdiction_coverage": "jurisdiction",
        "chief_sheriff":        "chief_name",
        "sworn_officers":       "sworn_officers",
        "headquarters_address": "headquarters",
        "phone":                "phone",
        "website":              "website",
    }
    records = _map_records(raw, col_map, "agency_name")
    print(f"  OK    Law Enforcement → law_enforcement.json ({len(records)} records)")
    return records


def extract_fire_ems(wb) -> list[dict]:
    ws = find_sheet(wb, "Fire")
    if ws is None:
        print("  WARN  sheet not found: Fire & EMS")
        return []
    raw = extract_records(ws)
    col_map = {
        "agency_type":              "agency_type",
        "department_authority_name": "name",
        "jurisdiction_area":        "jurisdiction",
        "fire_chief":               "chief_name",
        "stations":                 "stations",
        "personnel":                "personnel",
        "headquarters":             "headquarters",
        "phone":                    "phone",
        "website":                  "website",
        "service_type":             "service_type",
    }
    records = _map_records(raw, col_map, "department_authority_name")
    print(f"  OK    Fire & EMS → fire_ems.json ({len(records)} records)")
    return records


def extract_hospitals(wb) -> list[dict]:
    ws = find_sheet(wb, "Hospital")
    if ws is None:
        print("  WARN  sheet not found: Hospitals & Healthcare")
        return []
    raw = extract_records(ws)
    col_map = {
        "ownership_type":       "ownership_type",
        "hospital_facility_name": "name",
        "county":               "county",
        "city":                 "city",
        "beds":                 "beds",
        "trauma_level":         "trauma_level",
        "health_system":        "health_system",
        "ceo_administrator":    "ceo",
        "phone":                "phone",
        "website":              "website",
    }
    records = _map_records(raw, col_map, "hospital_facility_name")
    print(f"  OK    Hospitals & Healthcare → hospitals.json ({len(records)} records)")
    return records


def extract_utilities_transit(wb) -> list[dict]:
    ws = find_sheet(wb, "Utilities")
    if ws is None:
        print("  WARN  sheet not found: Utilities & Transit")
        return []
    raw = extract_records(ws)
    col_map = {
        "category":              "category",
        "agency_district_name":  "name",
        "county_region":         "county_region",
        "service_type":          "service_type",
        "customers_riders":      "customers_riders",
        "ceo_executive_director": "ceo",
        "phone":                 "phone",
        "website":               "website",
        "governing_board":       "governing_board",
    }
    records = _map_records(raw, col_map, "agency_district_name")
    print(f"  OK    Utilities & Transit → utilities_transit.json ({len(records)} records)")
    return records


def extract_judiciary(wb) -> list[dict]:
    ws = find_sheet(wb, "Judiciary")
    if ws is None:
        print("  WARN  sheet not found: Judiciary")
        return []
    raw = extract_records(ws)
    col_map = {
        "court_level":        "court_level",
        "court_name":         "court_name",
        "position":           "position",
        "justice_judge_name": "judge_name",
        "selection_method":   "selection_method",
        "appointed_by_elected": "appointed_by",
        "term_start":         "term_start",
        "term_end":           "term_end",
        "jurisdiction":       "jurisdiction",
        "official_website":   "website",
    }
    records = _map_records(raw, col_map, "justice_judge_name")
    print(f"  OK    Judiciary → judiciary.json ({len(records)} records)")
    return records


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    for sheet_name, output_file in SHEET_MAP.items():
        ws = find_sheet(wb, sheet_name)
        if ws is None:
            print(f"  WARN  sheet not found: {sheet_name}")
            continue

        records = extract_records(ws)
        out_path = OUTPUT_DIR / output_file

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2, ensure_ascii=False)

        print(f"  OK    {sheet_name} → {output_file} ({len(records)} records)")

    voting_records = extract_voting_records(wb)
    out_path = OUTPUT_DIR / "voting_records.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(voting_records, f, indent=2, ensure_ascii=False)
    print(f"  OK    voting records → voting_records.json ({len(voting_records)} total)")

    campaign_finance = extract_campaign_finance(wb)
    out_path = OUTPUT_DIR / "campaign_finance.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(campaign_finance, f, indent=2, ensure_ascii=False)
    print(f"  OK    campaign finance → campaign_finance.json ({len(campaign_finance)} total)")

    elections, election_candidates = extract_elections(wb)
    with open(OUTPUT_DIR / "elections.json", "w", encoding="utf-8") as f:
        json.dump(elections, f, indent=2, ensure_ascii=False)
    print(f"  OK    elections → elections.json ({len(elections)} total)")
    with open(OUTPUT_DIR / "candidates.json", "w", encoding="utf-8") as f:
        json.dump(election_candidates, f, indent=2, ensure_ascii=False)
    print(f"  OK    candidates → candidates.json ({len(election_candidates)} total)")

    for extractor, filename in [
        (extract_school_boards,    "school_boards.json"),
        (extract_state_agencies,   "state_agencies.json"),
        (extract_law_enforcement,  "law_enforcement.json"),
        (extract_fire_ems,         "fire_ems.json"),
        (extract_hospitals,        "hospitals.json"),
        (extract_utilities_transit, "utilities_transit.json"),
        (extract_judiciary,        "judiciary.json"),
    ]:
        records = extractor(wb)
        with open(OUTPUT_DIR / filename, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2, ensure_ascii=False)


if __name__ == "__main__":
    main()
